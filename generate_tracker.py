#!/usr/bin/env python3
"""
Issues & MAPs Tracker Generator

Reads from FIG Issue Management Model (repaired_2).xlsx and produces a
per-business-unit workbook with three sheets:
  - Issues and MAPs       : full BU issue/MAP data
  - MAPs Compliance Sheet : open/past-due MAPs with compliance metrics
  - Summary               : high-level metrics
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit this section before running
# ─────────────────────────────────────────────────────────────────────────────

# Directory where output workbooks are written
username = "" ## REPLACE WITH USERNAME
BASE_DIR = f"C:/Users/{username}/Downloads"

# Path to the source workbook (daily-updated)
SOURCE_FILE = f"{BASE_DIR}/FIG Issue Management Model (repaired_2).xlsx"

INPUT_FILE = "DPS-EPS - Issue and MAPs - 06162026.xlsx"

# Set to True to save a debug snapshot of the full merged dataset before
# any BU filtering, so you can spot duplicate MAP IDs or missing columns.
SAVE_DEBUG_SNAPSHOT = True
DEBUG_SNAPSHOT_PATH = f"{BASE_DIR}/debug_all_issues_and_maps.xlsx"

# ── Business-unit definitions ────────────────────────────────────────────────
# Add one dict per BU.  Fields:
#   business_unit_name    : used in the output filename
#   business_leader_names : values to match in col_to_search (list of strings)
#   col_to_search         : column in the merged output to filter on
#                           (use the renamed display name, e.g. "Issue MC-2")
#   previous_tracker_path : path to last week's output file for this BU,
#                           or None to skip comment carry-forward
BUSINESS_UNITS = [
    {
        "business_unit_name": "DPS-EPS",
        "business_leader_names": [],
        "col_to_search": "MC-3 Name",
        "previous_tracker_path": None,
    },
    {
        "business_unit_name": "CAPS",
        "business_leader_names": [],
        "col_to_search": "MC-3 Name",
        "previous_tracker_path": None,
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN DEFINITIONS — adjust if source column names differ
# ─────────────────────────────────────────────────────────────────────────────

# Columns to pull from "dump issues" (exact names as they appear in the sheet)
ISSUES_SOURCE_COLS = [
    "Issue ID",
    "MC -1",
    "MC -2",
    "MC -3",
    "Issue Name**",
    "Issue Source L1**",
    "Issue Workflow Status",
    "Date Opened",
    "Issue Due Date",
    "Enterprise Risk Severity Rating",
]

# Pre-merge rename map for issues columns (source name → display name).
# Applied before the join to prevent column-name collisions.
ISSUES_RENAME = {
    "MC -2": "Issue MC-2",
    "MC -1": "Issue MC-1",
    "MC -3": "MC-3 Name",
    "Issue Name**": "Issue Name",
    "Issue Source L1**": "Source",
    "Issue Workflow Status": "Issue Status",
    "Date Opened": "issue date opened",
}

# Columns to pull from "dump maps" (exact names as they appear in the sheet)
MAPS_SOURCE_COLS = [
    "Issue Id_Calc_Issue",
    "MAP ID",
    "MAP Name**",
    "MC -2",
    "MAP Workflow Status",
    "MAP Owner",
    "MAP Opened Date",
    "MAP Due Date (Current)",
    "AP Status",
    "AP Summary Status",
    "Last Updated",
    "# Days to MAP Due Date",
]

# Pre-merge rename map for maps columns
MAPS_RENAME = {
    "MAP Name**": "MAP Name",
    "MC -2": "MAP MC2",
    "MAP Workflow Status": "MAP Status",
    "MAP Opened Date": "MAP opened date",
    "MAP Due Date (Current)": "MAP Due Date",
    "AP Summary Status": "Summary Update",
    "Last Updated": "Last Updated Date",
}

# Ordered columns in the "Issues and MAPs" output sheet.
# "Comments", "If MAP is Past Due, ETA?", and "hash" are added by the script.
ISSUES_AND_MAPS_COLUMNS = [
    "Hash",
    "Issue MC-1",
    "Issue MC-2",
    "Issue ID",
    "Issue Name",
    "Source",
    "MC-3 Name",
    "Issue Status",
    "issue date opened",
    "Issue Due Date",
    "MAP MC2",
    "MAP Owner",
    "MAP Status",
    "MAP ID",
    "MAP Name",
    "MAP opened date",
    "MAP Due Date",
    "AP Status",
    "Enterprise Risk Severity Rating",
    "Last Updated Date",
    "# Days to MAP Due Date",
    "Summary Update",
    "Comments",
    "If MAP is Past Due, ETA?",
]

# Issue Status values that indicate a discussion is needed
ISSUE_STATUSES_INCLUDE = {"open", "past due", "past due - pending map owner approval", "past due - pending aso approval"}

# MAP Status values that indicate no discussion is needed
MAP_STATUSES_EXCLUDE = {"cancelled", "draft", "completed", "draft", "map cancelled", "draft - pending approvals", "map cancellation pending aso approval"}

# Columns shown on the "MAPs Compliance Sheet"
COMPLIANCE_DISPLAY_COLUMNS = [
    "Issue ID",
    "Issue Name",
    "MAP ID",
    "MAP Name",
    "MAP Status",
    "MAP Owner",
    "MAP Due Date",
    "Last Updated Date",
    "Days Since Update",
    "Cadence Days",
    "Compliance Result",
]

# Date columns — formatted as MM/DD/YYYY in the output sheets
DATE_COLUMNS = {
    "issue date opened",
    "Issue Due Date",
    "MAP opened date",
    "MAP Due Date",
    "Last Updated Date",
}

# ─────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=11)
_BOLD_FONT = Font(name="Calibri", bold=True, size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=13)
_COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_NONCOMPLIANT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_source_data(source_file: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load 'dump issues' and 'dump maps' from the source workbook."""
    path = Path(source_file)
    if not path.exists():
        print(f"ERROR: Source file not found: {path.resolve()}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading: {path.resolve()}")
    issues_df = pd.read_excel(source_file, sheet_name="dump issues", dtype=str)
    maps_df = pd.read_excel(source_file, sheet_name="dump maps", dtype=str)

    # Normalize column names (strip whitespace)
    issues_df.columns = issues_df.columns.str.strip()
    maps_df.columns = maps_df.columns.str.strip()

    print(f"  dump issues : {len(issues_df):,} rows, {len(issues_df.columns)} columns")
    print(f"  dump maps   : {len(maps_df):,} rows, {len(maps_df.columns)} columns")
    return issues_df, maps_df


def _select_available_cols(df: pd.DataFrame, wanted: list[str], label: str) -> pd.DataFrame:
    """Select columns that exist; warn about any that are missing."""
    missing = [c for c in wanted if c not in df.columns]
    if missing:
        print(f"  WARNING [{label}]: columns not found and will be skipped: {missing}")
    present = [c for c in wanted if c in df.columns]
    return df[present].copy()

# ─────────────────────────────────────────────────────────────────────────────
# MAP DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def get_latest_maps(maps_df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep only the most recent row per MAP ID, using 'Last Updated'.
    Each MAP in the returned frame has a unique MAP ID.
    """
    df = maps_df.copy()
    df["Last Updated"] = pd.to_datetime(df["Last Updated"], errors="coerce")
    latest = (
        df.sort_values("Last Updated", ascending=False)
        .drop_duplicates(subset=["MAP ID"], keep="first")
        .reset_index(drop=True)
    )
    print(f"  MAP dedup   : {len(maps_df):,} rows → {len(latest):,} unique MAP IDs")
    return latest

# ─────────────────────────────────────────────────────────────────────────────
# MERGE ISSUES + MAPS
# ─────────────────────────────────────────────────────────────────────────────

def build_all_issues_and_maps(
    issues_df: pd.DataFrame,
    maps_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Merge issues and maps, rename columns to display names, add empty
    user-maintained columns (Comments, ETA), and add the 'hash' key.
    """
    # Select and rename issue columns before the join
    issues = _select_available_cols(issues_df, ISSUES_SOURCE_COLS, "dump issues")
    issues = issues.rename(columns=ISSUES_RENAME)

    # Select and rename map columns before the join
    maps = _select_available_cols(maps_df, MAPS_SOURCE_COLS, "dump maps")
    maps = maps.rename(columns=MAPS_RENAME)

    # Right-join so every MAP appears even if its Issue ID is not found
    merged = issues.merge(
        maps,
        left_on="Issue ID",
        right_on="Issue Id_Calc_Issue",
        how="right",
    )
    merged = merged.drop(columns=["Issue Id_Calc_Issue"], errors="ignore")

    # Add user-maintained columns (populated later from previous tracker)
    merged["Comments"] = ""
    merged["If MAP is Past Due, ETA?"] = ""

    # Composite key for week-over-week comment carry-forward
    merged["Hash"] = (
        merged.get("Issue ID", pd.Series("", index=merged.index)).fillna("").astype(str)
        + "|"
        + merged.get("MAP ID", pd.Series("", index=merged.index)).fillna("").astype(str)
    )

    # Order to the desired output columns, keeping only what exists
    ordered = [c for c in ISSUES_AND_MAPS_COLUMNS if c in merged.columns]
    merged = merged[ordered].reset_index(drop=True)

    print(f"  Merged      : {len(merged):,} rows, {len(merged.columns)} columns")
    return merged


def save_debug_snapshot(df: pd.DataFrame, path: str) -> None:
    """Write the full merged dataset to an Excel file for manual spot-checking."""
    df.to_excel(path, index=False)
    print(f"  Debug snapshot saved → {Path(path).resolve()}")

# ─────────────────────────────────────────────────────────────────────────────
# DISCUSSION-NEEDED FILTER
# ─────────────────────────────────────────────────────────────────────────────

def filter_discussion_needed(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return only rows that need active discussion — i.e. where both the Issue
    Status and the MAP Status indicate in-flight work.

    Excluded (→ N) when either column is blank-like (NaN / empty string / 0)
    or matches a terminal status:
      Issue Status : closed, cancelled
      MAP Status   : cancelled, completed, draft
    """
    issue_status = df["Issue Status"]
    issue_active = (
        ~issue_status.isna() # not null
        & (issue_status.astype(str).str.strip() != "") # not empty
        & (issue_status != 0) # not blank/zero
        & issue_status.astype(str).str.strip().str.lower().isin(ISSUE_STATUSES_INCLUDE) # has these statuses
    )

    map_status = df["MAP Status"]
    map_active = (
        ~map_status.isna() # not null
        & (map_status.astype(str).str.strip() != "") # not empty
        & (map_status != 0) # not blank/zero
        & ~map_status.astype(str).str.strip().str.lower().isin(MAP_STATUSES_EXCLUDE) # not these statuses
    )

    result = df[issue_active & map_active].copy().reset_index(drop=True)
    excluded = len(df) - len(result)
    print(f"  Discussion-needed filter: {len(result):,} rows kept, {excluded:,} excluded")
    return result

# ─────────────────────────────────────────────────────────────────────────────
# BUSINESS-UNIT FILTERING
# ─────────────────────────────────────────────────────────────────────────────

def filter_by_business_unit(
    df: pd.DataFrame,
    col_to_search: str,
    business_leader_names: list[str],
) -> pd.DataFrame:
    """Return rows where col_to_search matches any value in business_leader_names."""
    if col_to_search not in df.columns:
        print(
            f"  WARNING: filter column '{col_to_search}' not found in data. "
            f"Available: {list(df.columns)}"
        )
        return df.iloc[0:0].copy()

    mask = df[col_to_search].isin(business_leader_names)
    result = df[mask].copy().reset_index(drop=True)
    print(f"  BU filter   : {len(result):,} rows matched {business_leader_names!r} in '{col_to_search}'")
    return result

# ─────────────────────────────────────────────────────────────────────────────
# COMMENT CARRY-FORWARD FROM PREVIOUS TRACKER
# ─────────────────────────────────────────────────────────────────────────────

def merge_data_from_tracker(
    current_df: pd.DataFrame,
    previous_tracker_path: str | None,
) -> pd.DataFrame:
    """
    Pull the 'Comments' and 'If MAP is Past Due, ETA?' columns from last
    week's output file and merge them into current_df by matching on 'hash'.
    New MAP IDs (no match) keep empty strings.
    """
    if not previous_tracker_path:
        print("  Previous tracker: none configured, skipping comment carry-forward.")
        return current_df

    prev_path = Path(previous_tracker_path)
    if not prev_path.exists():
        print(f"  Previous tracker: not found at {prev_path}, skipping.")
        return current_df

    print(f"  Merging comments from: {prev_path}")
    prev = pd.read_excel(prev_path, sheet_name="Issues and MAPs", dtype=str)
    prev.columns = prev.columns.str.strip()

    carry = ["hash", "Comments", "If MAP is Past Due, ETA?"]
    carry = [c for c in carry if c in prev.columns]

    if "hash" not in carry:
        print("  WARNING: 'hash' column not found in previous tracker; skipping merge.")
        return current_df

    prev_subset = prev[carry].rename(
        columns={
            "Comments": "_prev_Comments",
            "If MAP is Past Due, ETA?": "_prev_ETA",
        }
    )

    merged = current_df.merge(prev_subset, on="hash", how="left")

    if "_prev_Comments" in merged.columns:
        merged["Comments"] = merged["_prev_Comments"].fillna("")
        merged.drop(columns=["_prev_Comments"], inplace=True)

    if "_prev_ETA" in merged.columns:
        merged["If MAP is Past Due, ETA?"] = merged["_prev_ETA"].fillna("")
        merged.drop(columns=["_prev_ETA"], inplace=True)

    return merged

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY AND COMMENTS ENRICHMENT
# ─────────────────────────────────────────────────────────────────────────────

def enrich_summary_and_comments(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prefix 'Summary Update' with the 'Last Updated Date' month/day (e.g. 'Jan 11: ...'),
    and prefix 'Comments' with today's date (e.g. 'Jun - 18: \\n'), appending any
    existing comment text carried forward from the previous tracker after a newline.
    """
    df = df.copy()
    today_prefix = date.today().strftime("%b - %d: ")

    # Prefix Summary Update with Last Updated Date month and day
    if "Summary Update" in df.columns and "Last Updated Date" in df.columns:
        def _format_summary(row):
            summary = str(row["Summary Update"]) if pd.notna(row["Summary Update"]) else ""
            last_updated = row["Last Updated Date"]
            if pd.isna(last_updated) or str(last_updated).strip() in ("", "nan", "NaT"):
                return summary
            try:
                dt = pd.to_datetime(last_updated)
                date_prefix = dt.strftime("%b %d")
            except Exception:
                date_prefix = str(last_updated).strip()
            return f"{date_prefix}: {summary}" if summary else f"{date_prefix}: "

        df["Summary Update"] = df.apply(_format_summary, axis=1)

    # Prefix Comments with today's date; append old comments after a newline
    if "Comments" in df.columns:
        def _format_comments(comment):
            existing = str(comment).strip() if pd.notna(comment) and str(comment).strip() not in ("", "nan") else ""
            if existing:
                return f"{today_prefix}\n{existing}"
            return f"{today_prefix}\n"

        df["Comments"] = df["Comments"].apply(_format_comments)

    return df

# ─────────────────────────────────────────────────────────────────────────────
# MAP COMPLIANCE CALCULATION
# ─────────────────────────────────────────────────────────────────────────────

def generate_map_compliance_data(bu_df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter to Open / Past Due MAPs and compute:
      - Days Since Update  = today − Last Updated Date
      - Cadence Days       = 7 (Past Due) | 14 (≤60 days to due) | 30 (>60 days)
      - Compliance Result  = Compliant | Non-Compliant
    """
    today = pd.Timestamp(date.today())

    open_statuses = {"Open", "Past Due"}
    compliance_df = bu_df[
        bu_df["MAP Status"].isin(open_statuses)
    ].copy().reset_index(drop=True)

    # Parse dates
    compliance_df["Last Updated Date"] = pd.to_datetime(
        compliance_df["Last Updated Date"], errors="coerce"
    )
    compliance_df["Days Since Update"] = (
        (today - compliance_df["Last Updated Date"]).dt.days
        .fillna(-1)
        .astype(int)
    )

    days_to_due = pd.to_numeric(
        compliance_df.get("# Days to MAP Due Date", pd.Series(dtype=float)),
        errors="coerce",
    ).fillna(999)

    def _cadence(row_status: str, days_remaining: float) -> int:
        if row_status == "Past Due":
            return 7
        return 14 if days_remaining <= 60 else 30

    compliance_df["Cadence Days"] = [
        _cadence(status, days)
        for status, days in zip(compliance_df["MAP Status"], days_to_due)
    ]

    compliance_df["Compliance Result"] = compliance_df.apply(
        lambda r: "Compliant"
        if r["Days Since Update"] >= 0 and r["Days Since Update"] <= r["Cadence Days"]
        else "Non-Compliant",
        axis=1,
    )

    n_compliant = (compliance_df["Compliance Result"] == "Compliant").sum()
    n_noncompliant = (compliance_df["Compliance Result"] == "Non-Compliant").sum()
    print(
        f"  Compliance  : {len(compliance_df):,} open/past-due MAPs | "
        f"{n_compliant} compliant, {n_noncompliant} non-compliant"
    )
    return compliance_df

# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK WRITING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _write_df_to_sheet(
    ws,
    df: pd.DataFrame,
    date_cols: set[str] | None = None,
    drop_cols: list[str] | None = None,
) -> None:
    """Write a DataFrame to an openpyxl worksheet with styling."""
    date_cols = date_cols or set()
    display_df = df.drop(columns=[c for c in (drop_cols or []) if c in df.columns])

    # Header row
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    # Data rows
    for row_idx, row in enumerate(display_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            col_name = display_df.columns[col_idx - 1]
            if col_name in date_cols and pd.notna(value) and str(value).strip() not in ("", "NaT", "nan"):
                try:
                    parsed = pd.to_datetime(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=parsed.to_pydatetime())
                    cell.number_format = "MM/DD/YYYY"
                except Exception:
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            else:
                safe = "" if pd.isna(value) or str(value) in ("nan", "NaT") else str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=safe)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Auto-width (approximate, capped)
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        letter = get_column_letter(col_idx)
        width = min(max(len(str(col_name)), 10) + 2, 40)
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _color_compliance_column(ws, compliance_col_idx: int) -> None:
    """Apply red/green fill to cells in the Compliance Result column."""
    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=compliance_col_idx,
        max_col=compliance_col_idx,
    ):
        for cell in row:
            if cell.value == "Compliant":
                cell.fill = _COMPLIANT_FILL
            elif cell.value == "Non-Compliant":
                cell.fill = _NONCOMPLIANT_FILL


def _write_summary_sheet(
    ws,
    bu_df: pd.DataFrame,
    compliance_df: pd.DataFrame,
    bu_name: str,
) -> None:
    """Populate the Summary sheet with aggregate metrics."""
    today_str = date.today().strftime("%m/%d/%Y")
    total_maps = bu_df["MAP ID"].dropna().nunique() if "MAP ID" in bu_df.columns else 0
    n_open = len(compliance_df)
    n_compliant = int((compliance_df.get("Compliance Result", pd.Series()) == "Compliant").sum())
    n_noncompliant = int((compliance_df.get("Compliance Result", pd.Series()) == "Non-Compliant").sum())

    pct_compliant = f"{n_compliant / n_open * 100:.1f}%" if n_open > 0 else "N/A"
    pct_noncompliant = f"{n_noncompliant / n_open * 100:.1f}%" if n_open > 0 else "N/A"

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value=f"Issues & MAPs Tracker — {bu_name}")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _HEADER_FILL
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")

    rows = [
        ("Date Generated", today_str),
        None,
        ("Total MAPs (all statuses)", total_maps),
        ("Open / Past Due MAPs", n_open),
        ("Compliant MAPs", n_compliant),
        ("Non-Compliant MAPs", n_noncompliant),
        ("% Compliant", pct_compliant),
        ("% Non-Compliant", pct_noncompliant),
    ]

    for i, entry in enumerate(rows, start=3):
        if entry is None:
            continue
        label, value = entry
        lbl = ws.cell(row=i, column=1, value=label)
        val = ws.cell(row=i, column=2, value=value)
        lbl.font = _BOLD_FONT
        val.font = _BODY_FONT
        lbl.border = _THIN_BORDER
        val.border = _THIN_BORDER
        if label == "% Non-Compliant" and n_noncompliant > 0:
            val.fill = _NONCOMPLIANT_FILL
        elif label == "% Compliant" and n_compliant > 0:
            val.fill = _COMPLIANT_FILL

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.row_dimensions[1].height = 24

# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

def write_workbook(
    bu_df: pd.DataFrame,
    compliance_df: pd.DataFrame,
    business_unit_name: str,
    output_dir: str,
) -> str:
    """Write the three-sheet tracker workbook and return the output path."""
    today_str = date.today().strftime("%Y%m%d")
    filename = f"Issues & MAPs Tracker - {business_unit_name} - {today_str}.xlsx"
    output_path = Path(output_dir) / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Sheet 1: Issues and MAPs ─────────────────────────────────────────────
    ws_issues = wb.create_sheet("Issues and MAPs")
    _write_df_to_sheet(ws_issues, bu_df, date_cols=DATE_COLUMNS)

    # ── Sheet 2: MAPs Compliance Sheet ──────────────────────────────────────
    ws_compliance = wb.create_sheet("MAPs Compliance Sheet")
    compliance_display = compliance_df[
        [c for c in COMPLIANCE_DISPLAY_COLUMNS if c in compliance_df.columns]
    ]
    _write_df_to_sheet(ws_compliance, compliance_display, date_cols=DATE_COLUMNS)

    if "Compliance Result" in compliance_display.columns:
        comp_col_idx = list(compliance_display.columns).index("Compliance Result") + 1
        _color_compliance_column(ws_compliance, comp_col_idx)

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, bu_df, compliance_df, business_unit_name)

    wb.save(str(output_path))
    print(f"  Output      : {output_path.resolve()}")
    return str(output_path)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 65)
    print("  Issues & MAPs Tracker Generator")
    print("=" * 65)

    # 1. Load source sheets
    issues_df, maps_df = load_source_data(SOURCE_FILE)

    # 2. Deduplicate MAPs to most-recent row per MAP ID
    latest_maps_df = get_latest_maps(maps_df)

    # 3. Merge issues + maps into a single flat table
    print("\nBuilding merged dataset...")
    all_issues_and_maps = build_all_issues_and_maps(issues_df, latest_maps_df)

    # 4. Save debug snapshot for manual verification (optional)
    if SAVE_DEBUG_SNAPSHOT:
        print(f"\nSaving debug snapshot for manual review...")
        save_debug_snapshot(all_issues_and_maps, DEBUG_SNAPSHOT_PATH)
        print("  Review this file to confirm columns are correct and there are no duplicate MAP IDs.")

    # 5. Filter to rows that need active discussion (open issues + active MAPs)
    print("\nFiltering for open, past-due issues and maps with discussion needed...")
    all_issues_and_maps_with_discussion_needed = filter_discussion_needed(all_issues_and_maps)

    # 6. Generate one workbook per business unit
    print(f"\nProcessing {len(BUSINESS_UNITS)} business unit(s)...\n")
    for bu_config in BUSINESS_UNITS:
        bu_name = bu_config["business_unit_name"]
        leader_names = bu_config["business_leader_names"]
        col_to_search = bu_config.get("col_to_search", "Issue MC-2")
        prev_tracker = bu_config.get("previous_tracker_path")

        print(f"{'─' * 55}")
        print(f"  [{bu_name}]")

        bu_data = filter_by_business_unit(all_issues_and_maps_with_discussion_needed, col_to_search, leader_names)
        if bu_data.empty:
            print(f"  SKIP: no data found for '{bu_name}' — check business_leader_names and col_to_search.")
            continue

        bu_data = merge_data_from_tracker(bu_data, prev_tracker)
        bu_data = enrich_summary_and_comments(bu_data)
        compliance_data = generate_map_compliance_data(bu_data)
        write_workbook(bu_data, compliance_data, bu_name, BASE_DIR)

    print(f"\n{'=' * 65}")
    print("  Done.")
    print(f"{'=' * 65}\n")


if __name__ == "__main__":
    main()
